# -*- coding: cp1252 -*-
from openpyxl import load_workbook
import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.patches import Circle
from matplotlib.collections import PatchCollection
import numpy as np
import sys,math

def give_col(length,diam):
    cols = list(plt.rcParams['axes.prop_cycle'])
    if length<16:
        if diam<1:
            c = cols[0]
        else:
            c = cols[1]
    else:
        if diam<1:
            c = cols[2]
        else:
            c = cols[3]
    return c['color']

modele = 'v1'
variant = ''
wb = load_workbook(filename='../../../recu/2021.11.03 Pfahllayout V2/3257 FH Bern - Exportlasten_vh2, 2021.11.03.xlsx',data_only=True)

sheet = wb.worksheets[3]

pos = []
lab = []
diam = []
length = []
ldict = {70000:15,
         80000:20,
         90000:15,
         100000:20,
         120000:15,
         140000:20}
ddict = {70000:0.9,
         80000:0.9,
         120000:0.9,
         90000:0.9,
         100000:1.2,
         140000:1.2}
         
kr = 0
Kzlist = []
##of = open('piles_v1.txt','w')
for row in sheet.iter_rows():
    kr += 1
    if kr>12 and row[1].value:
        pt = np.array([row[3].value,-row[4].value])
        lab.append(row[1].value)
        pos.append(pt)
        Kz = row[14].value
        Kzlist.append(Kz)
##        Kzlist.add(Kz)
##        diam.append(ddict[Kz])
##        length.append(ldict[Kz])
        diam.append(row[27].value)
        length.append(row[28].value)
##        of.write('%s;%1.1f;%1.0f\n'%(row[1].value,ddict[Kz],ldict[Kz]))
##of.close()


matdict = {0.9:22,1.2:23}
y0 = 543.4

f = open('template.inp')
of = open('M1230_piles'+modele+'_'+variant+'.inp','w')
for line in iter(lambda: f.readline(), ""):
    if '.pob' in line:
        of.write(line)
        for kp in range(len(pos)):
            x = pos[kp][0]
            z = pos[kp][1]
            of.write('%i %1.12e %1.12e %1.12e 0\n'%(kp*2+1,x,y0,z))
            of.write('%i %1.12e %1.12e %1.12e 0\n'%(kp*2+2,x,y0-length[kp],z))
    elif '.pil' in line:
        of.write(line)
        dl = 1
        for kp in range(len(pos)):
            nSeg = int(length[kp]/dl)+1
            string1 = '%i %i PILE 1 %i %1.6f '%(kp+1,kp+1,nSeg,dl)
            string2 = '0.020000 0 0 1 1 14 %i %i\n'%(kp*2+1,kp*2+2)
            of.write(string1+string2)
            of.write('0\n')
            of.write('0  %1.6e  %1.6e  %1.6e  %1.6e  %1.6e  %1.6e\n'%(0,0,0,0,0,0))
            of.write('0  %1.6e  %1.6e  %1.6e  %1.6e  %1.6e  %1.6e\n'%(1,0,0,1,0,0))
            mat = matdict[diam[kp]]
            interfmat = 25
            tipmat = 26
            ef = 4
            of.write(' %i %i %i %i 0\n'%(mat,interfmat,ef,tipmat))
            x = pos[kp][0]
            z = pos[kp][1]
            y1 = y0-length[kp]
            for ks in range(nSeg+1):
                of.write('%1.12e %1.12e %1.12e 0\n'%(x,y0+(y1-y0)/nSeg*ks,z))
    else:
        of.write(line)
of.close()
f.close()

if True:
    pos = [np.array([v[0],-v[1]]) for v in pos]
    bbox_props = dict(boxstyle="round,pad=0", fc="w", ec="k", lw=0.4)
    fig = plt.figure(figsize=(16,8))
    ax = fig.add_subplot(111)
    patches = []
    cvect = []
    dscale = 0.8
    diams = sorted(list(set(diam)))
    lengths = sorted(list(set(length)))
    pilestat = [[0 for kl in range(len(lengths))] for kd in range(len(diams))]
    for kp in range(len(lab)):
        ax.add_patch(Circle(pos[kp],radius=0.5*diam[kp]/dscale,
                            fc=give_col(length[kp],diam[kp])))
##        cvect.append(length[kp]*diam[kp])
##        ax.annotate('%i'%(1e-4*Kzlist[kp]),#+':L%1.0f'%(length[kp]),
        ax.annotate('%i'%(lab[kp]),#+':L%1.0f'%(length[kp]),
                    xy=pos[kp],xytext=(2,-2),textcoords='offset points',#xytext=(pos[kp][0],pos[kp][1]),
                    ha='left',va='center',size=10,
                    bbox=bbox_props,rotation=20)
        pilestat[diams.index(diam[kp])][lengths.index(length[kp])] += 1

    cmap = plt.cm.get_cmap('Set1')
##    pc = PatchCollection(patches,edgecolors=('k',))
##    pc.set_array(np.array(cvect))
##    ax.add_collection(pc)

    cell_text = [['%i'%(pilestat[kd][kl]) for kl in range(len(lengths))] for kd in range(len(diams))]
    rows = ['D=%1.1fm'%(d)+' '*10 for d in diams]
    columns = ['L=%1.0fm'%(l) for l in lengths]
    cell_cols = [[give_col(15,0.9),give_col(20,0.9)],
                 [give_col(15,1.2),give_col(20,1.2)]]
    table = plt.table(cellText=cell_text,rowLabels=rows,colLabels=columns,
                      cellColours=cell_cols,
                      bbox=[0.8,0.02,0.15,0.05*len(diams)])
    table.set_fontsize(12)

##    ax.set_xticks([53.12+k*6.7 for k in range(7)])
##    ax.set_xticklabels(['A','B','C','D','E','F','G'])
##    ax.set_yticks([117.73-k*7.3 for k in range(7)])
##    ax.set_yticklabels([str(6+k*5) for k in range(7)])
    
    bounds = [[40,235],[-125,-40]]
    ax.set_title('Pfahlvariante %s, %s'%(modele,variant),size=28)
    ax.set_xlim(bounds[0])
    ax.set_ylim(bounds[1])
    ax.set_aspect('equal')
    
##    for x in ax.get_xticks():
##        ax.plot([x,x],bounds[1],'k:')
##    for y in ax.get_yticks():
##        ax.plot(bounds[0],[y,y],'k:')
        
    fig.tight_layout()
    fig.savefig('Pfahllayout_'+modele+'_'+variant)
    plt.close(fig)



