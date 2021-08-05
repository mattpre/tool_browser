# -*- coding: cp1252 -*-
import numpy,os,math,cmath
from numpy import linalg as la
from numpy import fft
import matplotlib as mpl
import matplotlib.pyplot as plt
import matplotlib.tri as mtri
import cPickle as pickle
from matplotlib.patches import Circle
from matplotlib.collections import PatchCollection
from openpyxl import load_workbook
from scipy import interpolate
import six
from matplotlib import colors

import sys
sys.path.append('D:/Mandats/python/zsoil_tools')
##sys.path.append('//JAN-09/Mandats/Mandats/python/zsoil_tools')
import zsoil_results as zr
import time


def posFE(x,y,z):
    p1 = numpy.array([x,y,z])
    # local to national coords:
    a = -95.*math.pi/180
    rot = numpy.array([[math.cos(a),0,math.sin(a)],
                       [0,1,0],
                       [-math.sin(a),0,math.cos(a)]])
    rotinv = rot.transpose()
    p0 = rotinv.dot(p1)
    delta = numpy.array([-684904,0,-256312])
    return (p0[0]-delta[0],p0[1],-p0[2]+delta[2])
def posFEinv(x,y,z):
    p0 = numpy.array([x-684904,y,z-256312])
    # national coords to local:
    a = -95.*math.pi/180
    rot = numpy.array([[math.cos(a),0,math.sin(a)],
                       [0,1,0],
                       [-math.sin(a),0,math.cos(a)]])
    p1 = rot.dot(p0)
    return (p1[0],p1[1],p1[2])

pathname = 'D:/Mandats/M990_Circle/technique/QP5-8/piles'
modele = '_2016-04-21'
wb = load_workbook(filename=pathname+'/4220 The Circle-Strukturdaten Pfählung H13+H15'+modele+'_mitStrati.xlsx')

sheets = [wb.get_sheet_by_name('H13+H15')]

labstr = ['C']
pos = []
lab = []
diam = []
load = []
length = []
strat = []
for ks in range(1):
    for row in sheets[ks].iter_rows():
        if not row[0].value==None:
            if row[0].value[0]=='C':
                lab.append(labstr[ks]+row[0].value[1:])
                pos.append((row[3].value,row[4].value))
                diam.append(float(row[1].value))
                load.append(row[6].value)
                strat.append([row[11+k].value for k in range(5)])
                length.append(row[2].value)


pathname = '//192.168.1.51/Mandats sur H RAID0/M990_Circle/QP58'
pblist = ['M990_QP58v2_1_2016-04-21_qsu410_knpile',
          'M990_QP58v2_1_2016-04-21_qsu410_knpileHigh',
          'M990_QP58v2_1_2016-04-21_qsu410_knTipHigh',
          'M990_QP58v2_1_2016-04-21_qsu410_knTipVHigh']
pblist = ['M990_QP58v2_1_2016-04-21_qsu410_knTipHigh']

for kf,prob in enumerate(pblist):
    figname = prob
    try:
        res = pickle.load(open(prob+'.p', "rb" ))
        print prob+' loaded'
        tsteps = res.out_steps
        tvect = [res.steps[kt].time for kt in tsteps]
    except:
        res = zr.zsoil_results(pathname,prob)
        res.read_rcf()
        res.read_his()

        tvect = [3,4]
        tsteps = []
        for kt,step in enumerate(res.steps):
            if step.time in tvect and step.sf==0 and step.conv_status==-1:
                tsteps.append(kt)
##        tsteps = [5,6]
                
        res.out_steps = tsteps
        res.read_dat()
    ##        res.read_s02()
        res.read_s00()
        res.read_s04()
        res.read_s07()
        pickle.dump(res, open(prob+'.p', "wb" ) ,pickle.HIGHEST_PROTOCOL)

    if res.steps[tsteps[0]].conv_status==-1:
        stat0 = u'convergé'
    else:
        stat0 = u'non convergé'
    if res.steps[tsteps[1]].conv_status==-1:
        stat1 = u'convergé'
    else:
        stat1 = u'non convergé'
    print u'Différence entre T=%1.1f (nStep=%i,%s) et T=%1.1f (nStep=%i,%s)'%(res.steps[tsteps[0]].time,tsteps[0],stat0,
                                                                             res.steps[tsteps[1]].time,tsteps[1],stat1)
    elist = []
    py = []
    pilecrds = []
    FN = []
    for ke in range(res.nBeams):
        if res.beam.mat[ke] in [12,13,14,15]:
            inel = res.beam.inel[ke]
            pcrd = (res.coords[0][inel[0]],-res.coords[2][inel[0]])
            tag = 0
            for kp in range(len(pilecrds)):
                if abs(pcrd[0]-pilecrds[kp][0])+abs(pcrd[1]-pilecrds[kp][1])<1e-6:
                    elist[kp].append(ke)
                    py[kp].append(0.5*(res.coords[1][inel[0]]+res.coords[1][inel[1]]))
                    tag = 1
                    break
            if tag==0:
                elist.append([ke])
                pilecrds.append(pcrd)
                py.append([0.5*(res.coords[1][inel[0]]+res.coords[1][inel[1]])])

    ###################################################
    # contact:
##    elist = []
    clist = []
    ctlist = []
    for kp in range(len(pilecrds)):
        pcrd = pilecrds[kp]
##        el = []
        cl = []
        ctl = -1
##        for ke in elist:
##            inel = res.beam.inel[ke]
##            x = res.coords[0][inel[0]-1]
##            z = res.coords[2][inel[0]-1]
##            if abs(pcrd[0]-x)+abs(pcrd[1]-z)<1e-3:
##                el.append(ke)
        for ke in range(res.nContacts):
            inel = res.cnt.inel[ke]
            x = res.coords[0][inel[0]-1]
            z = -res.coords[2][inel[0]-1]
            if res.cnt.type[ke]==2:
                if abs(pcrd[0]-x)+abs(pcrd[1]-z)<1e-3:
                    cl.append(ke)
            elif res.cnt.type[ke]==3:
                if abs(pcrd[0]-x)+abs(pcrd[1]-z)<1e-3:
                    ctl = ke
##        elist.append(el)
        clist.append(cl)
        ctlist.append(ctl)
    ###################################################

    kntop = []  # nodes of top of pile
    for kp in range(len(pilecrds)):
        kntop.append(res.beam.inel[elist[kp][0]][0])
    # get pile cap settlement:
    step0 = res.steps[tsteps[0]]
    step = res.steps[tsteps[1]]
    dy_cap = []
    for kp in range(len(pilecrds)):
        dy_cap.append(min(0,step.nodal.disp[1][kntop[kp]-1]-step0.nodal.disp[1][kntop[kp]-1]))
    

    # get normal forces of piles:
    for el in elist:
        fn = []
        for ke in el:
            fn.append(step.beam.force[0][ke]-step0.beam.force[0][ke])
        FN.append(fn)



    # match piles with results:
    pind = [-1 for k in pilecrds]
    for kp in range(len(pilecrds)):
        for kp2 in range(len(pos)):
            if abs(pilecrds[kp][0]-pos[kp2][0])+abs(pilecrds[kp][1]-pos[kp2][1])<0.5:
                pind[kp] = kp2
                break
        if pind[kp]==-1:
            print 'pos %1.1f,%1.1f not found'%pilecrds[kp]

    diam2 = [diam[k] for k in pind]
    
    # pile details:
    yref = 424.05
    if 1==1:
        
        cdict = dict({0:u'#D38D5F',1:u'#d7d4e8',2:u'#f6f8bf',3:u'#c9e3c2',
                      4:u'#f6d6ab',5:u'#e2c6dd'})
        layer_names = ['Deckschichten',
                       'Junge Schotter',
                       'Junge Seeablagerungen',
                       u'Moräne',
                       u'Ältere Schotter',
                       u'Ältere Seeablagerungen']
        nPilFig = 10
        for kfig in range(len(pilecrds)/nPilFig+1):
                          
            sizex = 18
            sizey = 7
            fig = plt.figure(figsize=(sizex,sizey))
            bbox_props = dict(boxstyle="round,pad=0.1", fc="w", ec="k", lw=0.4)
            fig.text(0.01,0.01,prob,size=8)

            ax = fig.add_subplot(111)

            scale = 0.00015
            x0 = 0
            legentries = []
            for kkel,el in enumerate(elist[kfig*nPilFig:(kfig+1)*nPilFig]):
##                plength = yref-res.coords[1][res.beam.inel[el[-1]][1]-1]
                kel = nPilFig*kfig+kkel
                plength = res.coords[1][kntop[kel]-1]-res.coords[1][res.beam.inel[el[-1]][1]-1]
                x0 += 2.0
                for ky in range(len(py[kel])):
                    inel = res.beam.inel[el[ky]]
                    ax.plot([x0,x0],
                            [res.coords[1][inel[0]-1],res.coords[1][inel[1]-1]],'k')
                    ax.plot([x0-FN[kel][ky]*scale,x0-FN[kel][ky]*scale],
                            [res.coords[1][inel[0]-1],res.coords[1][inel[1]-1]],'k')
                    if ky==0:
                        ax.plot([x0,x0-FN[kel][ky]*scale],
                                [res.coords[1][inel[0]-1],res.coords[1][inel[0]-1]],'k')
                    else:
                        ax.plot([x0-FN[kel][ky-1]*scale,x0-FN[kel][ky]*scale],
                                [res.coords[1][inel[0]-1],res.coords[1][inel[0]-1]],'k')
                ax.plot([x0-FN[kel][-1]*scale,x0],
                        [res.coords[1][inel[1]-1],res.coords[1][inel[1]-1]],'k')
                    
                ax.annotate('%ikN'%(min(FN[kel])),xy=(x0-min(FN[kel])*scale,py[kel][FN[kel].index(min(FN[kel]))]+0.5),
                            ha='center',va='bottom',size=10)
##                ax.annotate('%ikN'%(FN[kel][-1]),xy=(x0-FN[kel][-1]*scale,py[kel][FN[kel].index(FN[kel][-1])]-0.5),
##                            ha='center',va='top',size=10)
                ax.annotate('%ikPa'%(step.cnt.stress[2][ctlist[kel]][0]),xy=(x0-FN[kel][-1]*scale,py[kel][FN[kel].index(FN[kel][-1])]-0.5),
                            ha='center',va='top',size=10)
                ax.annotate(lab[pind[kel]]+'\nL=%im\nD=%1.1fm'%(plength+0.1,diam2[kel]),
                            xy=(x0,max(py[kel])),xytext=(x0,max(py[kel])+2),ha='center',va='bottom',size=10)

                # geologie:
##            if 1==1:
                top = []
                lind = []
                for ks,s in enumerate(strat[pind[kel]]):
                    if not s==-1:# and s<yref:
                        if s>yref:
                            top.append(yref)
                        else:
                            top.append(s)
                        lind.append(ks+1)
                if not abs(top[0]-yref)<0.01:
                    top.insert(0,yref)
                    lind.insert(0,0)
                bot = [y for y in top[1:]]
##                bot.append(min(py[kel])-5)
                bot.append(350)
                height = [top[k]-bot[k] for k in range(len(top))]
##                lind.append(5)
                for kb in range(len(top)):
                    if lind[kb] in legentries:
                        ax.bar(x0-0.4,-height[kb],bottom=top[kb],color=cdict[lind[kb]])
                    else:
                        ax.bar(x0-0.4,-height[kb],bottom=top[kb],color=cdict[lind[kb]],label=layer_names[lind[kb]])
                        legentries.append(lind[kb])

            ax.legend(loc='lower right')
            ax.grid(b=True,which='both',axis='y')
            ax.set_ylim([345,435])
            fig.tight_layout()
            fig.savefig('pile_details/'+prob+'_piles_details_%i'%(kfig+1))
            fig.savefig('pile_details/'+prob+'_piles_details_%i.svg'%(kfig+1))
            plt.close(fig)


    bounds = [[min([crd[0] for crd in pilecrds])-2,max([crd[0] for crd in pilecrds])+2],
              [min([crd[1] for crd in pilecrds])-2,max([crd[1] for crd in pilecrds])+2]]
    # contour of setzungen:
##    get nodes of Bodenplatte
    nlist = set()
    for ke in range(res.nShells):
        for kn in res.shell.inel[ke]:
            nlist.add(kn-1)
    nlist = list(nlist)
    x = []
    z = []
    v = []
    for kn in nlist:
        v.append(min(0,res.steps[tsteps[-1]].nodal.disp[1][kn]-res.steps[tsteps[0]].nodal.disp[1][kn]))
        x.append(res.coords[0][kn])
        z.append(-res.coords[2][kn])
    tri = mtri.Triangulation(x,z)
    mask = mtri.TriAnalyzer(tri).get_flat_tri_mask(0.1)
    tri.set_mask(mask)
    f = mtri.LinearTriInterpolator(tri,v)
##    f = interpolate.Rbf(x,z,v,function='linear')
##    f = interpolate.Rbf(x,z,v,function='linear')
    xnew = numpy.arange(bounds[0][0],bounds[0][1],0.5)
    znew = numpy.arange(bounds[1][0],bounds[1][1],0.5)
    X,Z = numpy.meshgrid(xnew,znew)
    V = numpy.zeros([len(X),len(X[0])])
    for kx in range(len(X[0])):
        for kz in range(len(X)):
##            V[kz][kx] = interpolate.griddata((x,z),v,(X[kz][kx],Z[kz][kx]),method='linear')
            V[kz][kx] = f(X[kz][kx],Z[kz][kx])*1000
   
    fig = plt.figure(figsize=(18,22))
    bbox_props = dict(boxstyle="round,pad=0.1", fc="w", ec="k", lw=0.4)
    fig.text(0.01,0.01,prob,size=8)

    ax = fig.add_subplot(111)
    # boreholes:
    pts_bh = []
    labels = []
    f = open('D:/Mandats/M990_Circle/technique/geologiemodell/boreholes.txt')
    for line in iter(lambda: f.readline(), ""):
        if 'KB' in line:
            labels.append(line[:-1])
            line = f.readline()
            v = line.split()
            pts_bh.append(posFEinv(float(v[1]),0,-float(v[3])))
            ax.plot(pts_bh[-1][0],pts_bh[-1][2],'k.')
            ax.annotate(labels[-1],xy=(pts_bh[-1][0],pts_bh[-1][2]))
    # profilspuren:
    pts_ps = [[684886.05,256077.2],
              [684912.01,256021.49],
              [684959.49, 255972.16],
              [685028.45, 255952.52]]
    dirs_ps = [[-138.02,-31.65],
               [-116.44,-72.16],
               [-79.72,-105.69],
               [-32.29,-93.69]]
    v0 = -1
    v1 = 2

    for kpt in range(len(pts_ps)):
        pt0 = posFEinv(pts_ps[kpt][0]+v0*dirs_ps[kpt][0],0,(pts_ps[kpt][1]+v0*dirs_ps[kpt][1]))
        pt1 = posFEinv(pts_ps[kpt][0]+v1*dirs_ps[kpt][0],0,(pts_ps[kpt][1]+v1*dirs_ps[kpt][1]))
        ax.plot([pt0[0],pt1[0]],
                [pt0[2],pt1[2]],'k--',linewidth=2)
    #####
    patches = []
    dscale = 1
    bbox_props = dict(boxstyle="round,pad=0.1", fc="w", ec="k", lw=0.4)
    bbox_props = dict(boxstyle="round,pad=0", fc="w", ec="k", lw=0.4)
    for k in range(len(pilecrds)):
        el = elist[k]
##        plength = yref-res.coords[1][res.beam.inel[el[-1]][1]-1]
        plength = res.coords[1][kntop[k]-1]-res.coords[1][res.beam.inel[el[-1]][1]-1]
        patches.append(Circle(pilecrds[k],radius=diam2[k]*0.5/dscale))
    ##    patches.append(Circle(pos[k],radius=0.2/dscale))
        ax.annotate(lab[pind[k]]+',%ikN\nL%im,%1.0fmm'%(load[pind[k]],plength+0.1,dy_cap[k]*1000),xy=pilecrds[k],xytext=(pilecrds[k][0],pilecrds[k][1]-2.0),
                    ha='left',va='bottom',size=9,
                    bbox=bbox_props,rotation=45)
    ax.annotate(u'Pfahlnummer, Gebrauchslast\nLänge,Setzung Pfahlkopf',xy=(bounds[0][0]+7,bounds[1][1]-17),
                ha='left',va='bottom',size=12,
                bbox=bbox_props,rotation=45)

    cmap = plt.cm.get_cmap('jet')
    pc = PatchCollection(patches,edgecolors=('k',),cmap=cmap)
    pc.set_clim(vmin=min(diam2),vmax=max(diam2))
    pc.set_array(numpy.array(diam2))
    ax.add_collection(pc)

    ds = 2.0
    levels = [int(min(dy_cap)*1000)+k*ds for k in range(int(1000*(max(dy_cap)-min(dy_cap))/ds)+1)]
    CS = ax.contour(X,Z,V,levels)
##    CS = ax.contour(X,Z,V,N,linewidths=[(2,) for k in range(N)])
    ax.clabel(CS,inline=0,inline_spacing=2,fontsize=10,colors='k',fmt='%1.1f')

    for i in range(len(levels)):
        CS.collections[i].set_label('%1.1f mm'%(levels[i]))
    ax.legend(loc='lower right',title='Setzungen')
    
    ax.set_xticks([198.996+k*13.5 for k in range(-3,16)])
    ax.set_xticklabels(['BD','BI','BN','BS','BX','CC','CH','CM','CR','CW',
                        'DB','DG','DL','DQ','DV','EA','EF','EK','EP'])
    ax.set_yticks([6.731+k*13.5 for k in range(26)])
    ax.set_yticklabels([str(0+k*5) for k in range(26)])
    
    ax.set_xlim(bounds[0])
    ax.set_ylim(bounds[1])
    ax.set_aspect('equal')
    
    for x in ax.get_xticks():
        ax.plot([x,x],bounds[1],'k:')
    for y in ax.get_yticks():
        ax.plot(bounds[0],[y,y],'k:')

    fig.tight_layout()
    fig.savefig(prob+'_pileres')
    plt.close(fig)



##########################################################################
##    Excel Tabelle:
##if 1==1:
    of = open(prob+'_resultate.csv','w')
    string = 'Nummer;Länge;Durchmesser;Last;N_max;Setzung Pfahlkopf;qp;qs_mean;X;Z\n'
    of.write(string)
    string = ';[m];[cm];[kN];[kN];[mm];[kPa];[kPa];[m];[m]\n' 
    of.write(string)

    for k in range(len(pilecrds)):
        el = elist[k]
##        plength = yref-res.coords[1][res.beam.inel[el[-1]][1]-1]
        plength = res.coords[1][kntop[k]-1]-res.coords[1][res.beam.inel[el[-1]][1]-1]

        string = lab[pind[k]]
        string += ';%1.0f;%1.0f;%1.0f;%1.0f'%(plength,diam2[k]*100,load[pind[k]],-min(FN[k]))
        l = [abs(res.coords[1][res.cnt.inel[ke][0]-1]-res.coords[1][res.cnt.inel[ke][1]-1]) for ke in clist[k]]
        qsmean = sum([0.5*sum(step.cnt.stress[0][ke])*l[kk] for kk,ke in enumerate(clist[k])])/sum(l)
        string += ';%1.1f;%1.2f;%1.2f'%(dy_cap[k]*1000,step.cnt.stress[2][ctlist[k]][0],qsmean)
        string += ';%1.2f;%1.2f\n'%(pilecrds[k][0],pilecrds[k][1])
        of.write(string)
    of.close()







